"""
Depot Blueprint
Handles price list uploads and depot operations
"""

from flask import Blueprint, request, jsonify, session as flask_session
from app.database.central_db import CentralDB
from app.database.tenant_db import TenantDBManager
from app.services.file_processor import FileProcessor
from app.services.ocr_processor import get_ocr_processor
from app.utils.auth import require_auth, require_role
from sqlalchemy import text
import uuid
import os
from datetime import datetime
import logging
import traceback
import re
import json
from difflib import SequenceMatcher

logger = logging.getLogger(__name__)

depot_bp = Blueprint('depot', __name__)


@depot_bp.route('/upload', methods=['POST'])
@require_auth
@require_role('depot')
def upload_price_list():
    """Upload and process price list file"""
    try:
        # Get user info
        user_id = request.current_user['user_id']
        tenant_id = request.current_user.get('tenant_id')
        
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        
        tenant_db_name = tenant.database_name
        session.close()
        
        # Check file upload
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        filename = file.filename
        if not any(filename.lower().endswith(ext) for ext in ['.csv', '.xlsx', '.xls']):
            return jsonify({'error': 'Invalid file type. Only CSV and Excel files are allowed'}), 400
        
        # Read file content
        file_content = file.read()
        file_size = len(file_content)
        
        # Process file
        processor = FileProcessor()
        valid_records, valid_count, invalid_count = processor.process_file(file_content, filename)
        
        if valid_count == 0:
            return jsonify({
                'error': 'No valid records found',
                'errors': processor.errors[:10]  # First 10 errors
            }), 400
        
        # Save to tenant database
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        try:
            # Create price list record
            price_list_id = uuid.uuid4()
            # Convert errors dict to JSON string for JSONB column
            errors_json = None
            if processor.errors:
                errors_dict = {'errors': processor.errors[:50]}
                errors_json = json.dumps(errors_dict)
            
            tenant_session.execute(
                text("""
                    INSERT INTO price_lists (id, tenant_id, version, status, file_name, 
                        total_items, valid_items, invalid_items, processing_errors)
                    VALUES (:id, :tenant_id, 1, 'processing', :file_name,
                        :total_items, :valid_items, :invalid_items, CAST(:errors AS jsonb))
                """),
                {
                    'id': price_list_id,
                    'tenant_id': tenant_id,
                    'file_name': filename,
                    'total_items': valid_count + invalid_count,
                    'valid_items': valid_count,
                    'invalid_items': invalid_count,
                    'errors': errors_json
                }
            )
            
            # Insert/Update medicines - check by name across ALL price lists
            # If medicine exists anywhere, update it; otherwise insert new
            # Handle duplicates in the same file by using the first occurrence
            updated_count = 0
            inserted_count = 0
            processed_medicines = {}  # Track medicines already processed in this upload
            
            for record in valid_records:
                medicine_name = record['medicine_name'].strip()
                unit_price = record['unit_price']
                expiry_date = record.get('expiry_date')
                
                # Validate expiry_date is present
                if not expiry_date:
                    logger.warning(f"Row missing expiry_date for '{medicine_name}' - skipping")
                    continue
                
                # Normalize medicine name for duplicate detection
                medicine_key = medicine_name.lower().strip()
                
                # If we've already processed this medicine in this file, skip it (use first occurrence)
                if medicine_key in processed_medicines:
                    logger.info(f"Skipping duplicate '{medicine_name}' in upload (already processed with price {processed_medicines[medicine_key]})")
                    continue
                
                # Mark as processed
                processed_medicines[medicine_key] = unit_price
                
                # Check if medicine with this name exists anywhere in the database
                check_result = tenant_session.execute(
                    text("""
                        SELECT id, medicine_name, unit_price, expiry_date FROM medicines 
                        WHERE LOWER(TRIM(medicine_name)) = LOWER(TRIM(:medicine_name))
                        AND price_list_id IN (
                            SELECT id FROM price_lists WHERE tenant_id = :tenant_id
                        )
                        LIMIT 1
                    """),
                    {
                        'medicine_name': medicine_name,
                        'tenant_id': tenant_id
                    }
                )
                existing_medicine = check_result.fetchone()
                
                if existing_medicine:
                    # Medicine exists - update ALL instances of this medicine name with new price and expiry_date
                    # Use exact same comparison as check query (TRIM on both sides)
                    old_price = float(existing_medicine.unit_price) if existing_medicine.unit_price else 0
                    update_result = tenant_session.execute(
                        text("""
                            UPDATE medicines
                            SET unit_price = :unit_price,
                                expiry_date = :expiry_date,
                                last_updated = CURRENT_TIMESTAMP,
                                is_active = TRUE
                            WHERE LOWER(TRIM(medicine_name)) = LOWER(TRIM(:medicine_name))
                            AND price_list_id IN (
                                SELECT id FROM price_lists WHERE tenant_id = :tenant_id
                            )
                        """),
                        {
                            'medicine_name': medicine_name,
                            'unit_price': unit_price,
                            'expiry_date': expiry_date,
                            'tenant_id': tenant_id
                        }
                    )
                    rows_updated = update_result.rowcount
                    if rows_updated > 0:
                        updated_count += rows_updated
                        logger.info(f"Updated {rows_updated} instance(s) of medicine '{medicine_name}' from price {old_price} to {unit_price}, expiry_date: {expiry_date}")
                    else:
                        logger.warning(f"Medicine '{medicine_name}' found but UPDATE affected 0 rows - check query logic")
                else:
                    # Medicine doesn't exist - insert new
                    tenant_session.execute(
                        text("""
                            INSERT INTO medicines (price_list_id, medicine_name, unit_price, expiry_date)
                            VALUES (:price_list_id, :medicine_name, :unit_price, :expiry_date)
                        """),
                        {
                            'price_list_id': price_list_id,
                            'medicine_name': medicine_name,
                            'unit_price': unit_price,
                            'expiry_date': expiry_date
                        }
                    )
                    inserted_count += 1
                    logger.info(f"Inserted new medicine '{medicine_name}' with price {unit_price}, expiry_date: {expiry_date}")
            
            logger.info(f"Upload summary: {updated_count} updated, {inserted_count} inserted, {len(processed_medicines)} unique medicines processed")
            
            # Update price list status
            tenant_session.execute(
                text("""
                    UPDATE price_lists
                    SET status = 'active', activated_at = CURRENT_TIMESTAMP
                    WHERE id = :id
                """),
                {'id': price_list_id}
            )
            
            # Archive old price lists (but keep all medicines - don't delete them)
            # This just marks old price lists as archived, medicines remain visible
            tenant_session.execute(
                text("""
                    UPDATE price_lists
                    SET status = 'archived', archived_at = CURRENT_TIMESTAMP
                    WHERE tenant_id = :tenant_id AND id != :current_id AND status = 'active'
                """),
                {'tenant_id': tenant_id, 'current_id': price_list_id}
            )
            
            # Log upload history
            tenant_session.execute(
                text("""
                    INSERT INTO upload_history (user_id, file_name, file_size_bytes,
                        records_processed, records_success, records_failed, status)
                    VALUES (:user_id, :file_name, :file_size, :total, :success, :failed, 'completed')
                """),
                {
                    'user_id': user_id,
                    'file_name': filename,
                    'file_size': file_size,
                    'total': valid_count + invalid_count,
                    'success': valid_count,
                    'failed': invalid_count
                }
            )
            
            tenant_session.commit()
            tenant_session.close()
            
            return jsonify({
                'message': 'Price list uploaded successfully',
                'upload_id': str(price_list_id),
                'statistics': {
                    'total_items': valid_count + invalid_count,
                    'valid_items': valid_count,
                    'invalid_items': invalid_count
                },
                'warnings': processor.warnings[:10] if processor.warnings else []
            }), 201
            
        except Exception as e:
            tenant_session.rollback()
            tenant_session.close()
            logger.error(f"Database error during upload: {e}")
            raise
        
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f"Upload error: {e}")
        return jsonify({'error': 'Upload failed', 'message': str(e)}), 500


@depot_bp.route('/upload-ocr', methods=['POST'])
@require_auth
@require_role('depot')
def upload_ocr_scan():
    """Scan image/PDF with OCR and return editable data for preview"""
    try:
        # Check file upload
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        filename = file.filename.lower()
        allowed_extensions = ['.png', '.jpg', '.jpeg', '.pdf', '.bmp', '.tiff', '.webp']
        if not any(filename.endswith(ext) for ext in allowed_extensions):
            return jsonify({
                'error': f'Invalid file type. Allowed: {", ".join(allowed_extensions)}'
            }), 400
        
        # Read file content
        file_content = file.read()
        
        # Initialize progress tracking in session
        progress_id = f"ocr_progress_{request.current_user.get('user_id')}"
        flask_session[progress_id] = {'percentage': 0, 'message': 'Starting...'}
        
        # Progress callback function
        def update_progress(percentage, message):
            flask_session[progress_id] = {'percentage': percentage, 'message': message}
            flask_session.modified = True
        
        # Process with OCR (use singleton for faster performance)
        ocr_processor = get_ocr_processor()
        result = ocr_processor.process_file(file_content, filename, progress_callback=update_progress)
        
        # Clear progress from session
        if progress_id in flask_session:
            del flask_session[progress_id]
        
        # Return structured data for editing
        return jsonify({
            'success': True,
            'text': result.get('text', ''),
            'structured_data': result.get('structured_data', []),
            'pages': result.get('pages', 1),
            'message': f'Successfully scanned {result.get("pages", 1)} page(s)'
        }), 200
        
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        logger.error(f"OCR scan error: {e}")
        return jsonify({'error': 'OCR scan failed', 'message': str(e)}), 500


@depot_bp.route('/upload-ocr-progress', methods=['GET'])
@require_auth
@require_role('depot')
def upload_ocr_progress():
    """Get current OCR scanning progress"""
    try:
        progress_id = f"ocr_progress_{request.current_user.get('user_id')}"
        progress = flask_session.get(progress_id, {'percentage': 0, 'message': 'Not started'})
        return jsonify(progress), 200
    except Exception as e:
        logger.error(f"Progress check error: {e}")
        return jsonify({'percentage': 0, 'message': 'Error checking progress'}), 500


@depot_bp.route('/upload-ocr-confirm', methods=['POST'])
@require_auth
@require_role('depot')
def upload_ocr_confirm():
    """Confirm and save OCR-scanned data after user edits"""
    try:
        user_id = request.current_user['user_id']
        tenant_id = request.current_user.get('tenant_id')
        
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Get edited data from request
        data = request.get_json()
        if not data or 'records' not in data:
            return jsonify({'error': 'No data provided'}), 400
        
        records = data['records']
        if not isinstance(records, list) or len(records) == 0:
            return jsonify({'error': 'No valid records provided'}), 400
        
        # Validate and process records (require medicine_name, unit_price, and expiry_date)
        valid_records = []
        for record in records:
            medicine_name = record.get('medicine_name', '').strip()
            unit_price = record.get('unit_price')
            expiry_date = record.get('expiry_date')
            
            if not medicine_name:
                continue
            
            if not expiry_date:
                continue  # Skip records without expiry_date
            
            try:
                price = float(unit_price) if unit_price else None
                if price and price <= 0:
                    continue
                
                # Validate expiry_date format
                from datetime import datetime
                try:
                    # Try to parse the date
                    if isinstance(expiry_date, str):
                        parsed_date = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                    else:
                        parsed_date = expiry_date
                    expiry_date_str = parsed_date.isoformat() if hasattr(parsed_date, 'isoformat') else str(parsed_date)
                except (ValueError, TypeError, AttributeError):
                    continue  # Skip invalid dates
                
                valid_records.append({
                    'medicine_name': medicine_name,
                    'unit_price': price if price else 0.0,
                    'expiry_date': expiry_date_str
                })
            except (ValueError, TypeError):
                continue
        
        if len(valid_records) == 0:
            return jsonify({'error': 'No valid records found'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        
        tenant_db_name = tenant.database_name
        session.close()
        
        # Save to tenant database (same logic as regular upload)
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        try:
            # Create price list record
            price_list_id = uuid.uuid4()
            tenant_session.execute(
                text("""
                    INSERT INTO price_lists (id, tenant_id, version, status, file_name, 
                        total_items, valid_items, invalid_items)
                    VALUES (:id, :tenant_id, 1, 'processing', :file_name,
                        :total_items, :valid_items, :invalid_items)
                """),
                {
                    'id': price_list_id,
                    'tenant_id': tenant_id,
                    'file_name': 'OCR_Scan',
                    'total_items': len(valid_records),
                    'valid_items': len(valid_records),
                    'invalid_items': 0
                }
            )
            
            # Insert/Update medicines - check by name across ALL price lists
            # If medicine exists anywhere, update it; otherwise insert new
            # Handle duplicates in the same upload by using the first occurrence
            processed_medicines = {}  # Track medicines already processed in this upload
            
            for record in valid_records:
                medicine_name = record['medicine_name'].strip()
                unit_price = record['unit_price']
                expiry_date = record.get('expiry_date')
                
                # Validate expiry_date is present
                if not expiry_date:
                    logger.warning(f"Row missing expiry_date for '{medicine_name}' - skipping")
                    continue
                
                # Normalize medicine name for duplicate detection
                medicine_key = medicine_name.lower().strip()
                
                # If we've already processed this medicine in this upload, skip it (use first occurrence)
                if medicine_key in processed_medicines:
                    logger.info(f"Skipping duplicate '{medicine_name}' in upload (already processed with price {processed_medicines[medicine_key]})")
                    continue
                
                # Mark as processed
                processed_medicines[medicine_key] = unit_price
                
                # Check if medicine with this name exists anywhere in the database
                check_result = tenant_session.execute(
                    text("""
                        SELECT id, medicine_name, unit_price, expiry_date FROM medicines 
                        WHERE LOWER(TRIM(medicine_name)) = LOWER(TRIM(:medicine_name))
                        AND price_list_id IN (
                            SELECT id FROM price_lists WHERE tenant_id = :tenant_id
                        )
                        LIMIT 1
                    """),
                    {
                        'medicine_name': medicine_name,
                        'tenant_id': tenant_id
                    }
                )
                existing_medicine = check_result.fetchone()
                
                if existing_medicine:
                    # Medicine exists - update ALL instances of this medicine name with new price and expiry_date
                    # Use exact same comparison as check query (TRIM on both sides)
                    old_price = float(existing_medicine.unit_price) if existing_medicine.unit_price else 0
                    update_result = tenant_session.execute(
                        text("""
                            UPDATE medicines
                            SET unit_price = :unit_price,
                                expiry_date = :expiry_date,
                                last_updated = CURRENT_TIMESTAMP,
                                is_active = TRUE
                            WHERE LOWER(TRIM(medicine_name)) = LOWER(TRIM(:medicine_name))
                            AND price_list_id IN (
                                SELECT id FROM price_lists WHERE tenant_id = :tenant_id
                            )
                        """),
                        {
                            'medicine_name': medicine_name,
                            'unit_price': unit_price,
                            'expiry_date': expiry_date,
                            'tenant_id': tenant_id
                        }
                    )
                    rows_updated = update_result.rowcount
                    if rows_updated > 0:
                        logger.info(f"Updated {rows_updated} instance(s) of medicine '{medicine_name}' from price {old_price} to {unit_price}, expiry_date: {expiry_date}")
                    else:
                        logger.warning(f"Medicine '{medicine_name}' found but UPDATE affected 0 rows")
                else:
                    # Medicine doesn't exist - insert new
                    tenant_session.execute(
                        text("""
                            INSERT INTO medicines (price_list_id, medicine_name, unit_price, expiry_date)
                            VALUES (:price_list_id, :medicine_name, :unit_price, :expiry_date)
                        """),
                        {
                            'price_list_id': price_list_id,
                            'medicine_name': medicine_name,
                            'unit_price': unit_price,
                            'expiry_date': expiry_date
                        }
                    )
                    logger.info(f"Inserted new medicine '{medicine_name}' with price {unit_price}, expiry_date: {expiry_date}")
            
            # Update price list status
            tenant_session.execute(
                text("""
                    UPDATE price_lists
                    SET status = 'active', activated_at = CURRENT_TIMESTAMP
                    WHERE id = :id
                """),
                {'id': price_list_id}
            )
            
            # Archive old price lists
            tenant_session.execute(
                text("""
                    UPDATE price_lists
                    SET status = 'archived', archived_at = CURRENT_TIMESTAMP
                    WHERE tenant_id = :tenant_id AND id != :current_id AND status = 'active'
                """),
                {'tenant_id': tenant_id, 'current_id': price_list_id}
            )
            
            # Log upload history
            tenant_session.execute(
                text("""
                    INSERT INTO upload_history (user_id, file_name, file_size_bytes,
                        records_processed, records_success, records_failed, status)
                    VALUES (:user_id, :file_name, :file_size, :total, :success, :failed, 'completed')
                """),
                {
                    'user_id': user_id,
                    'file_name': 'OCR_Scan',
                    'file_size': 0,
                    'total': len(valid_records),
                    'success': len(valid_records),
                    'failed': 0
                }
            )
            
            tenant_session.commit()
            tenant_session.close()
            
            return jsonify({
                'message': 'Price list uploaded successfully',
                'upload_id': str(price_list_id),
                'statistics': {
                    'total_items': len(valid_records),
                    'valid_items': len(valid_records),
                    'invalid_items': 0
                }
            }), 201
            
        except Exception as e:
            tenant_session.rollback()
            tenant_session.close()
            logger.error(f"Database error during OCR upload: {e}")
            raise
        
    except Exception as e:
        logger.error(f"OCR confirm upload error: {e}")
        return jsonify({'error': 'Upload failed', 'message': str(e)}), 500


@depot_bp.route('/download-prices-ocr', methods=['POST'])
@require_auth
@require_role('depot')
def download_prices_ocr():
    """Scan image/PDF with OCR to extract medicine names for price lookup using fuzzy matching"""
    try:
        # Check file upload
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        filename = file.filename.lower()
        allowed_extensions = ['.png', '.jpg', '.jpeg', '.pdf', '.bmp', '.tiff', '.webp']
        if not any(filename.endswith(ext) for ext in allowed_extensions):
            return jsonify({
                'error': f'Invalid file type. Allowed: {", ".join(allowed_extensions)}'
            }), 400
        
        # Read file content
        file_content = file.read()
        
        # Initialize progress tracking in session
        progress_id = f"ocr_progress_{request.current_user.get('user_id')}"
        flask_session[progress_id] = {'percentage': 0, 'message': 'Starting...'}
        
        # Progress callback function
        def update_progress(percentage, message):
            flask_session[progress_id] = {'percentage': percentage, 'message': message}
            flask_session.modified = True
        
        # Process with OCR (use singleton for faster performance)
        ocr_processor = get_ocr_processor()
        result = ocr_processor.process_file(file_content, filename, progress_callback=update_progress)
        
        # Clear progress from session
        if progress_id in flask_session:
            del flask_session[progress_id]
        
        # Extract medicine names from structured data (EXACT same logic as upload-ocr)
        # The structured_data from OCR processor already contains parsed medicine names and prices
        # We use the same extraction method as upload-ocr endpoint
        structured_data = result.get('structured_data', [])
        medicine_names = []
        
        # Extract medicine names from structured data (same as upload-ocr uses)
        for item in structured_data:
            # Each item should have 'medicine_name' field (and optionally 'unit_price')
            name = item.get('medicine_name', '').strip()
            if name and len(name) > 0:
                medicine_names.append(name)
        
        # If no structured data found, try to extract from plain text (fallback - same as upload-ocr)
        if not medicine_names:
            ocr_text = result.get('text', '')
            if ocr_text:
                lines = ocr_text.split('\n')
                for line in lines:
                    line = line.strip()
                    if line and len(line) > 2:  # Basic validation - skip very short lines
                        # Try to parse line using the same parse_line method
                        try:
                            parsed = ocr_processor.parse_line(line)
                            if parsed and parsed.get('medicine_name'):
                                name = parsed['medicine_name'].strip()
                                if name:
                                    medicine_names.append(name)
                        except Exception as e:
                            logger.debug(f"Failed to parse line '{line}': {e}")
                            # If parsing fails, try to use the line itself if it looks like a medicine name
                            # (contains letters and is not just numbers/symbols)
                            if any(c.isalpha() for c in line) and len(line) > 2:
                                medicine_names.append(line)
        
        # Remove duplicates while preserving order
        seen = set()
        unique_medicine_names = []
        for name in medicine_names:
            name_lower = name.lower().strip()
            if name_lower and name_lower not in seen:
                seen.add(name_lower)
                unique_medicine_names.append(name)
        
        medicine_names = unique_medicine_names
        
        if not medicine_names:
            return jsonify({
                'error': 'No medicine names found in the scanned document',
                'text': result.get('text', ''),
                'structured_data': structured_data,
                'debug_info': {
                    'structured_data_count': len(structured_data),
                    'text_length': len(result.get('text', ''))
                }
            }), 400
        
        # Now lookup prices in database (this is the only difference from upload)
        tenant_id = request.current_user.get('tenant_id')
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        
        tenant_db_name = tenant.database_name
        session.close()
        
        # Query medicines from tenant database
        tenant_session = TenantDBManager.get_session(tenant_db_name)

        # Load all active medicines once for fuzzy matching
        all_medicines_result = tenant_session.execute(
            text("""
                SELECT m.id, m.medicine_name, m.unit_price, m.expiry_date, pl.activated_at
                FROM medicines m
                INNER JOIN price_lists pl ON m.price_list_id = pl.id
                WHERE pl.tenant_id = :tenant_id 
                    AND m.is_active = TRUE
                ORDER BY pl.activated_at DESC
            """),
            {'tenant_id': tenant_id}
        )
        all_medicines = all_medicines_result.fetchall()

        def _normalize_name(name_str: str) -> str:
            """Lowercase, remove numbers and most punctuation, collapse spaces."""
            if not name_str:
                return ''
            name_str = name_str.lower()
            # Remove digits
            name_str = re.sub(r'\d+', ' ', name_str)
            # Replace punctuation with space but keep letters and spaces
            name_str = re.sub(r'[^a-z\s]', ' ', name_str)
            # Collapse multiple spaces
            return ' '.join(name_str.split())

        matched_medicines = []
        not_found = []
        similarity_threshold = 0.70

        for name in medicine_names:
            try:
                name_clean = name.strip()
                if not name_clean:
                    continue

                target_normalized = _normalize_name(name_clean)
                if not target_normalized:
                    not_found.append(name_clean)
                    continue

                best_match = None
                best_score = 0.0

                for med in all_medicines:
                    med_normalized = _normalize_name(med.medicine_name)
                    if not med_normalized:
                        continue

                    # --- 1) Strong prefix rule: if first 3–4 letters match, force a high score ---
                    # Remove spaces for prefix comparison
                    target_no_space = target_normalized.replace(' ', '')
                    med_no_space = med_normalized.replace(' ', '')
                    max_prefix = min(len(target_no_space), len(med_no_space), 6)
                    prefix_len = 0
                    while prefix_len < max_prefix and target_no_space[prefix_len] == med_no_space[prefix_len]:
                        prefix_len += 1

                    if prefix_len >= 4:
                        score = 0.99  # almost perfect if first 4+ letters match
                    elif prefix_len >= 3:
                        score = 0.9   # good match if first 3 letters match
                    else:
                        # --- 2) Fallback to fuzzy similarity ---
                        score_char = SequenceMatcher(None, target_normalized, med_normalized).ratio()
                        target_tokens = ' '.join(sorted(target_normalized.split()))
                        med_tokens = ' '.join(sorted(med_normalized.split()))
                        score_tokens = SequenceMatcher(None, target_tokens, med_tokens).ratio()
                        score = max(score_char, score_tokens)

                    if score > best_score:
                        best_score = score
                        best_match = med

                if best_match and best_score >= similarity_threshold:
                    expiry_date_str = best_match.expiry_date.isoformat() if best_match.expiry_date else None
                    matched_medicines.append({
                        'id': str(best_match.id),
                        'scanned_name': name_clean,
                        'medicine_name': best_match.medicine_name,
                        'unit_price': float(best_match.unit_price),
                        'expiry_date': expiry_date_str,
                        'match_score': round(best_score, 3)
                    })
                else:
                    not_found.append(name_clean)
            except Exception as e:
                logger.error(f"Error querying medicine '{name}': {e}")
                not_found.append(name)

        tenant_session.close()
        
        return jsonify({
            'success': True,
            'matched_medicines': matched_medicines,
            'not_found': not_found,
            'total_requested': len(medicine_names),
            'total_matched': len(matched_medicines),
            'total_not_found': len(not_found),
            'text': result.get('text', ''),
            'structured_data': result.get('structured_data', []),
            'pages': result.get('pages', 1)
        }), 200
        
    except ValueError as e:
        logger.error(f"Download prices OCR ValueError: {e}", exc_info=True)
        return jsonify({'error': str(e), 'message': str(e)}), 400
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"Download prices OCR error: {e}")
        logger.error(f"Full traceback:\n{error_details}")
        # Return more detailed error for debugging
        return jsonify({
            'error': 'OCR scan failed', 
            'message': str(e),
            'error_type': type(e).__name__
        }), 500


@depot_bp.route('/download-prices-ocr-progress', methods=['GET'])
@require_auth
@require_role('depot')
def download_prices_ocr_progress():
    """Get current OCR scanning progress for download prices"""
    try:
        progress_id = f"ocr_progress_{request.current_user.get('user_id')}"
        progress = flask_session.get(progress_id, {'percentage': 0, 'message': 'Not started'})
        return jsonify(progress), 200
    except Exception as e:
        logger.error(f"Progress check error: {e}")
        return jsonify({'percentage': 0, 'message': 'Error checking progress'}), 500


@depot_bp.route('/price-lists', methods=['GET'])
@require_auth
@require_role('depot')
def get_price_lists():
    """Get all price lists for the depot"""
    try:
        tenant_id = request.current_user.get('tenant_id')
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        session.close()
        
        tenant_db_name = tenant.database_name
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        status = request.args.get('status', 'active')
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 10))
        offset = (page - 1) * limit
        
        result = tenant_session.execute(
            text("""
                SELECT id, version, status, file_name, total_items, valid_items,
                    invalid_items, activated_at, created_at
                FROM price_lists
                WHERE tenant_id = :tenant_id AND status = :status
                ORDER BY created_at DESC
                LIMIT :limit OFFSET :offset
            """),
            {'tenant_id': tenant_id, 'status': status, 'limit': limit, 'offset': offset}
        )
        price_lists = result.fetchall()
        
        # Get total count
        count_result = tenant_session.execute(
            text("SELECT COUNT(*) FROM price_lists WHERE tenant_id = :tenant_id AND status = :status"),
            {'tenant_id': tenant_id, 'status': status}
        )
        total = count_result.scalar()
        
        tenant_session.close()
        
        return jsonify({
            'price_lists': [{
                'id': str(pl.id),
                'version': pl.version,
                'status': pl.status,
                'file_name': pl.file_name,
                'total_items': pl.total_items,
                'valid_items': pl.valid_items,
                'invalid_items': pl.invalid_items,
                'activated_at': pl.activated_at.isoformat() if pl.activated_at else None,
                'created_at': pl.created_at.isoformat() if pl.created_at else None
            } for pl in price_lists],
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total,
                'pages': (total + limit - 1) // limit
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Get price lists error: {e}")
        return jsonify({'error': 'Failed to get price lists', 'message': str(e)}), 500


@depot_bp.route('/medicines', methods=['GET'])
@require_auth
@require_role('depot')
def get_medicines():
    """Get all medicines from active price list"""
    try:
        tenant_id = request.current_user.get('tenant_id')
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        session.close()
        
        tenant_db_name = tenant.database_name
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        # Get search query
        search = request.args.get('search', '').strip()
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 100))
        offset = (page - 1) * limit
        
        # Build query - show ALL medicines from ALL price lists (name, price, and expiry_date)
        if search:
            query = text("""
                SELECT m.id, m.medicine_name, m.unit_price, m.expiry_date, m.is_active,
                    pl.file_name, pl.status as price_list_status, pl.activated_at
                FROM medicines m
                INNER JOIN price_lists pl ON m.price_list_id = pl.id
                WHERE pl.tenant_id = :tenant_id 
                    AND m.medicine_name ILIKE :search
                ORDER BY pl.activated_at DESC, m.medicine_name
                LIMIT :limit OFFSET :offset
            """)
            params = {
                'tenant_id': tenant_id,
                'search': f'%{search}%',
                'limit': limit,
                'offset': offset
            }
        else:
            query = text("""
                SELECT m.id, m.medicine_name, m.unit_price, m.expiry_date, m.is_active,
                    pl.file_name, pl.status as price_list_status, pl.activated_at
                FROM medicines m
                INNER JOIN price_lists pl ON m.price_list_id = pl.id
                WHERE pl.tenant_id = :tenant_id 
                ORDER BY pl.activated_at DESC, m.medicine_name
                LIMIT :limit OFFSET :offset
            """)
            params = {
                'tenant_id': tenant_id,
                'limit': limit,
                'offset': offset
            }
        
        result = tenant_session.execute(query, params)
        medicines = result.fetchall()
        
        # Get total count - count ALL medicines from ALL price lists (including inactive)
        # This gives the true total of all medicines ever uploaded
        if search:
            count_query = text("""
                SELECT COUNT(DISTINCT m.id)
                FROM medicines m
                INNER JOIN price_lists pl ON m.price_list_id = pl.id
                WHERE pl.tenant_id = :tenant_id 
                    AND m.medicine_name ILIKE :search
            """)
        else:
            count_query = text("""
                SELECT COUNT(DISTINCT m.id)
                FROM medicines m
                INNER JOIN price_lists pl ON m.price_list_id = pl.id
                WHERE pl.tenant_id = :tenant_id
            """)
        
        count_result = tenant_session.execute(
            count_query,
            {'tenant_id': tenant_id, 'search': f'%{search}%'} if search else {'tenant_id': tenant_id}
        )
        total = count_result.scalar()
        
        tenant_session.close()
        
        return jsonify({
            'medicines': [{
                'id': str(m.id),
                'medicine_name': m.medicine_name,
                'unit_price': float(m.unit_price),
                'expiry_date': m.expiry_date.isoformat() if m.expiry_date else None,
                'is_active': m.is_active,
                'price_list_file': m.file_name if hasattr(m, 'file_name') else None,
                'price_list_status': m.price_list_status if hasattr(m, 'price_list_status') else None
            } for m in medicines],
            'pagination': {
                'page': page,
                'limit': limit,
                'total': total,
                'pages': (total + limit - 1) // limit
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Get medicines error: {e}")
        return jsonify({'error': 'Failed to get medicines', 'message': str(e)}), 500


@depot_bp.route('/upload-history', methods=['GET'])
@require_auth
@require_role('depot')
def get_upload_history():
    """Get upload history for the depot"""
    try:
        tenant_id = request.current_user.get('tenant_id')
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        session.close()
        
        tenant_db_name = tenant.database_name
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        result = tenant_session.execute(
            text("""
                SELECT id, file_name, file_size_bytes, upload_timestamp,
                    records_processed, records_success, records_failed, status
                FROM upload_history
                ORDER BY upload_timestamp DESC
                LIMIT 50
            """)
        )
        history = result.fetchall()
        tenant_session.close()
        
        return jsonify({
            'history': [{
                'id': str(h.id),
                'file_name': h.file_name,
                'file_size': h.file_size_bytes,
                'upload_timestamp': h.upload_timestamp.isoformat() if h.upload_timestamp else None,
                'records_processed': h.records_processed,
                'records_success': h.records_success,
                'records_failed': h.records_failed,
                'status': h.status
            } for h in history]
        }), 200
        
    except Exception as e:
        logger.error(f"Get upload history error: {e}")
        return jsonify({'error': 'Failed to get upload history', 'message': str(e)}), 500


@depot_bp.route('/upload-history/<history_id>', methods=['DELETE'])
@require_auth
@require_role('depot')
def delete_upload_history(history_id):
    """Delete an upload history entry"""
    try:
        tenant_id = request.current_user.get('tenant_id')
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        session.close()
        
        tenant_db_name = tenant.database_name
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        # Check if history entry exists
        hist_result = tenant_session.execute(
            text("SELECT id, file_name FROM upload_history WHERE id = :id"),
            {'id': history_id}
        )
        history = hist_result.fetchone()
        
        if not history:
            tenant_session.close()
            return jsonify({'error': 'History entry not found'}), 404
        
        # Delete history entry
        tenant_session.execute(
            text("DELETE FROM upload_history WHERE id = :id"),
            {'id': history_id}
        )
        
        tenant_session.commit()
        tenant_session.close()
        
        return jsonify({
            'message': 'History entry deleted successfully',
            'history': {
                'id': str(history.id),
                'file_name': history.file_name
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Delete upload history error: {e}")
        return jsonify({'error': 'Failed to delete history entry', 'message': str(e)}), 500


@depot_bp.route('/statistics', methods=['GET'])
@require_auth
@require_role('depot')
def get_statistics():
    """Get depot statistics"""
    try:
        tenant_id = request.current_user.get('tenant_id')
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        session.close()
        
        tenant_db_name = tenant.database_name
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        # Get statistics
        stats_result = tenant_session.execute(
            text("""
                SELECT 
                    COUNT(DISTINCT m.id) as total_medicines,
                    COUNT(DISTINCT pl.id) as total_uploads,
                    COUNT(DISTINCT CASE WHEN pl.status = 'active' THEN pl.id END) as active_price_lists
                FROM medicines m
                LEFT JOIN price_lists pl ON m.price_list_id = pl.id
                WHERE pl.tenant_id = :tenant_id
            """),
            {'tenant_id': tenant_id}
        )
        stats = stats_result.fetchone()
        
        tenant_session.close()
        
        return jsonify({
            'total_medicines': stats.total_medicines or 0,
            'total_uploads': stats.total_uploads or 0,
            'active_price_list': stats.active_price_lists or 0
        }), 200
        
    except Exception as e:
        logger.error(f"Get statistics error: {e}")
        return jsonify({'error': 'Failed to get statistics', 'message': str(e)}), 500


@depot_bp.route('/medicine/<medicine_id>', methods=['PUT'])
@require_auth
@require_role('depot')
def update_medicine(medicine_id):
    """Update a medicine"""
    try:
        tenant_id = request.current_user.get('tenant_id')
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        data = request.get_json()
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        session.close()
        
        tenant_db_name = tenant.database_name
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        # Build update query (only name and price allowed)
        updates = {}
        allowed_fields = ['medicine_name', 'unit_price', 'is_active']
        for field in allowed_fields:
            if field in data:
                updates[field] = data[field]
        
        if not updates:
            tenant_session.close()
            return jsonify({'error': 'No valid fields to update'}), 400
        
        # Update medicine
        set_clause = ', '.join([f"{k} = :{k}" for k in updates.keys()])
        updates['id'] = medicine_id
        
        result = tenant_session.execute(
            text(f"""
                UPDATE medicines
                SET {set_clause}, last_updated = CURRENT_TIMESTAMP
                WHERE id = :id
                RETURNING id, medicine_name, unit_price
            """),
            updates
        )
        medicine = result.fetchone()
        
        if not medicine:
            tenant_session.close()
            return jsonify({'error': 'Medicine not found'}), 404
        
        tenant_session.commit()
        tenant_session.close()
        
        return jsonify({
            'message': 'Medicine updated successfully',
            'medicine': {
                'id': str(medicine.id),
                'medicine_name': medicine.medicine_name,
                'unit_price': float(medicine.unit_price)
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Update medicine error: {e}")
        return jsonify({'error': 'Failed to update medicine', 'message': str(e)}), 500


@depot_bp.route('/medicine/<medicine_id>', methods=['DELETE'])
@require_auth
@require_role('depot')
def delete_medicine(medicine_id):
    """Delete a medicine"""
    try:
        tenant_id = request.current_user.get('tenant_id')
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        session.close()
        
        tenant_db_name = tenant.database_name
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        # Check if medicine exists
        med_result = tenant_session.execute(
            text("SELECT id, medicine_name FROM medicines WHERE id = :id"),
            {'id': medicine_id}
        )
        medicine = med_result.fetchone()
        
        if not medicine:
            tenant_session.close()
            return jsonify({'error': 'Medicine not found'}), 404
        
        # Delete medicine
        tenant_session.execute(
            text("DELETE FROM medicines WHERE id = :id"),
            {'id': medicine_id}
        )
        
        tenant_session.commit()
        tenant_session.close()
        
        return jsonify({
            'message': 'Medicine deleted successfully',
            'medicine': {
                'id': str(medicine.id),
                'medicine_name': medicine.medicine_name
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Delete medicine error: {e}")
        return jsonify({'error': 'Failed to delete medicine', 'message': str(e)}), 500


@depot_bp.route('/price-list/<price_list_id>', methods=['DELETE'])
@require_auth
@require_role('depot')
def delete_price_list(price_list_id):
    """Delete a price list and its medicines"""
    try:
        tenant_id = request.current_user.get('tenant_id')
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        session.close()
        
        tenant_db_name = tenant.database_name
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        # Check if price list exists
        pl_result = tenant_session.execute(
            text("SELECT id, file_name FROM price_lists WHERE id = :id AND tenant_id = :tenant_id"),
            {'id': price_list_id, 'tenant_id': tenant_id}
        )
        price_list = pl_result.fetchone()
        
        if not price_list:
            tenant_session.close()
            return jsonify({'error': 'Price list not found'}), 404
        
        # Delete medicines first (cascade)
        tenant_session.execute(
            text("DELETE FROM medicines WHERE price_list_id = :id"),
            {'id': price_list_id}
        )
        
        # Delete price list
        tenant_session.execute(
            text("DELETE FROM price_lists WHERE id = :id"),
            {'id': price_list_id}
        )
        
        tenant_session.commit()
        tenant_session.close()
        
        return jsonify({
            'message': 'Price list deleted successfully',
            'price_list': {
                'id': str(price_list.id),
                'file_name': price_list.file_name
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Delete price list error: {e}")
        return jsonify({'error': 'Failed to delete price list', 'message': str(e)}), 500


@depot_bp.route('/download-prices', methods=['POST'])
@require_auth
@require_role('depot')
def download_prices():
    """Process file with medicine names and return matched medicines with prices"""
    try:
        tenant_id = request.current_user.get('tenant_id')
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Check file upload
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        filename = file.filename.lower()
        if not filename.endswith(('.csv', '.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file type. Only CSV, XLSX, and XLS files are allowed'}), 400
        
        # Read file content
        file_content = file.read()
        
        # Extract medicine names from file
        file_processor = FileProcessor()
        df = None
        try:
            if filename.endswith('.csv'):
                df = file_processor._read_csv(file_content)
            elif filename.endswith(('.xlsx', '.xls')):
                df = file_processor._read_excel(file_content)
            else:
                return jsonify({'error': 'Unsupported file type'}), 400
        except Exception as e:
            logger.error(f"File read error: {e}")
            return jsonify({'error': f'Failed to read file: {str(e)}'}), 400
        
        if df is None or df.empty:
            return jsonify({'error': 'File is empty or could not be read'}), 400
        
        # Find medicine name column (case-insensitive)
        medicine_name_col = None
        for col in df.columns:
            col_lower = str(col).lower().strip()
            if any(keyword in col_lower for keyword in ['medicine', 'name', 'product', 'drug', 'item']):
                medicine_name_col = col
                break
        
        if medicine_name_col is None:
            # Try first column if no match found
            medicine_name_col = df.columns[0]
        
        # Extract medicine names
        medicine_names = df[medicine_name_col].dropna().astype(str).str.strip()
        medicine_names = medicine_names[medicine_names != ''].unique().tolist()
        
        if not medicine_names:
            return jsonify({'error': 'No medicine names found in file'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        
        tenant_db_name = tenant.database_name
        session.close()
        
        # Query medicines from database
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        # Build query to match medicine names (case-insensitive)
        matched_medicines = []
        not_found_names = []
        
        for name in medicine_names:
            name_clean = name.strip()
            if not name_clean:
                continue
            
            # First try exact match (case-insensitive)
            result = tenant_session.execute(
                text("""
                    SELECT m.id, m.medicine_name, m.unit_price
                    FROM medicines m
                    INNER JOIN price_lists pl ON m.price_list_id = pl.id
                    WHERE pl.tenant_id = :tenant_id 
                        AND LOWER(TRIM(m.medicine_name)) = LOWER(TRIM(:name))
                        AND m.is_active = TRUE
                    ORDER BY pl.activated_at DESC
                    LIMIT 1
                """),
                {'tenant_id': tenant_id, 'name': name_clean}
            )
            medicine = result.fetchone()
            
            # If no exact match, try matching without numbers
            if not medicine:
                # Remove numbers and normalize for matching
                name_normalized = re.sub(r'\d+', '', name_clean.lower().strip())
                name_normalized = ' '.join(name_normalized.split())  # Normalize spaces
                
                # Get all medicines and match manually
                all_medicines_result = tenant_session.execute(
                    text("""
                        SELECT m.id, m.medicine_name, m.unit_price, pl.activated_at
                        FROM medicines m
                        INNER JOIN price_lists pl ON m.price_list_id = pl.id
                        WHERE pl.tenant_id = :tenant_id 
                            AND m.is_active = TRUE
                        ORDER BY pl.activated_at DESC
                    """),
                    {'tenant_id': tenant_id}
                )
                all_medicines = all_medicines_result.fetchall()
                
                # Find best match (medicine name without numbers matches)
                best_match = None
                for med in all_medicines:
                    med_name_normalized = re.sub(r'\d+', '', med.medicine_name.lower().strip())
                    med_name_normalized = ' '.join(med_name_normalized.split())
                    
                    # Check if normalized names match
                    if name_normalized == med_name_normalized:
                        best_match = med
                        break
                    # Also check if input is contained in medicine name (without numbers)
                    elif name_normalized in med_name_normalized or med_name_normalized in name_normalized:
                        if not best_match:  # Take first match
                            best_match = med
                
                if best_match:
                    medicine = best_match
            
            if medicine:
                matched_medicines.append({
                    'id': str(medicine.id),
                    'scanned_name': name_clean,
                    'medicine_name': medicine.medicine_name,
                    'unit_price': float(medicine.unit_price)
                })
            else:
                not_found_names.append(name_clean)
        
        tenant_session.close()
        
        # Update medicines table to only contain names and prices
        # This means we'll keep the existing structure but ensure we only return name and price
        # The user wants the table to be simplified, so we'll update the query to only select name and price
        
        return jsonify({
            'message': 'Prices retrieved successfully',
            'matched_medicines': matched_medicines,
            'not_found': not_found_names,
            'statistics': {
                'total_requested': len(medicine_names),
                'matched': len(matched_medicines),
                'not_found': len(not_found_names)
            }
        }), 200
        
    except Exception as e:
        logger.error(f"Download prices error: {e}", exc_info=True)
        return jsonify({'error': 'Failed to process file', 'message': str(e)}), 500


@depot_bp.route('/download-prices-excel', methods=['POST'])
@require_auth
@require_role('depot')
def download_prices_excel():
    """Generate and download Excel file with medicine prices"""
    try:
        from flask import send_file
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        import io
        
        data = request.get_json()
        medicines = data.get('medicines', [])
        
        if not medicines:
            return jsonify({'error': 'No medicines data provided'}), 400
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Medicine Prices"
        
        # Add headers
        headers = ['Medicine Name', 'Unit Price', 'Expiry Date']
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add data
        for med in medicines:
            expiry_date = med.get('expiry_date', '')
            ws.append([
                med.get('medicine_name', ''),
                med.get('unit_price', 0),
                expiry_date
            ])
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        # Format price column
        from openpyxl.styles.numbers import BUILTIN_FORMATS
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = BUILTIN_FORMATS[4]  # Currency format
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'medicine_prices_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        logger.error(f"Excel download error: {e}", exc_info=True)
        return jsonify({'error': 'Failed to generate Excel file', 'message': str(e)}), 500


@depot_bp.route('/upload-manual', methods=['POST'])
@require_auth
@require_role('depot')
def upload_manual():
    """Upload prices via manual entry table"""
    try:
        user_id = request.current_user['user_id']
        tenant_id = request.current_user.get('tenant_id')
        
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Get data from request
        data = request.get_json()
        if not data or 'records' not in data:
            return jsonify({'error': 'No data provided'}), 400
        
        records = data['records']
        if not isinstance(records, list) or len(records) == 0:
            return jsonify({'error': 'No valid records provided'}), 400
        
        # Validate and process records (require medicine_name, unit_price, and expiry_date)
        valid_records = []
        for record in records:
            medicine_name = record.get('medicine_name', '').strip()
            unit_price = record.get('unit_price')
            expiry_date = record.get('expiry_date')
            
            if not medicine_name:
                continue
            
            if not expiry_date:
                continue  # Skip records without expiry_date
            
            try:
                price = float(unit_price) if unit_price else None
                if price and price <= 0:
                    continue
                
                # Validate expiry_date format
                from datetime import datetime
                try:
                    # Try to parse the date
                    if isinstance(expiry_date, str):
                        parsed_date = datetime.strptime(expiry_date, '%Y-%m-%d').date()
                    else:
                        parsed_date = expiry_date
                    expiry_date_str = parsed_date.isoformat() if hasattr(parsed_date, 'isoformat') else str(parsed_date)
                except (ValueError, TypeError, AttributeError):
                    continue  # Skip invalid dates
                
                valid_records.append({
                    'medicine_name': medicine_name,
                    'unit_price': price if price else 0.0,
                    'expiry_date': expiry_date_str
                })
            except (ValueError, TypeError):
                continue
        
        if len(valid_records) == 0:
            return jsonify({'error': 'No valid records found'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        
        tenant_db_name = tenant.database_name
        session.close()
        
        # Save to tenant database (same logic as OCR confirm)
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        try:
            # Create price list record
            price_list_id = uuid.uuid4()
            tenant_session.execute(
                text("""
                    INSERT INTO price_lists (id, tenant_id, version, status, file_name, 
                        total_items, valid_items, invalid_items)
                    VALUES (:id, :tenant_id, 1, 'processing', :file_name,
                        :total_items, :valid_items, :invalid_items)
                """),
                {
                    'id': price_list_id,
                    'tenant_id': tenant_id,
                    'file_name': 'manual_entry',
                    'total_items': len(valid_records),
                    'valid_items': len(valid_records),
                    'invalid_items': 0
                }
            )
            
            # Insert/Update medicines - check by name across ALL price lists
            # If medicine exists anywhere, update it; otherwise insert new
            # Handle duplicates in the same upload by using the first occurrence
            processed_medicines = {}  # Track medicines already processed in this upload
            
            for record in valid_records:
                medicine_name = record['medicine_name'].strip()
                unit_price = record['unit_price']
                expiry_date = record.get('expiry_date')
                
                # Validate expiry_date is present
                if not expiry_date:
                    logger.warning(f"Row missing expiry_date for '{medicine_name}' - skipping")
                    continue
                
                # Normalize medicine name for duplicate detection
                medicine_key = medicine_name.lower().strip()
                
                # If we've already processed this medicine in this upload, skip it (use first occurrence)
                if medicine_key in processed_medicines:
                    logger.info(f"Skipping duplicate '{medicine_name}' in upload (already processed with price {processed_medicines[medicine_key]})")
                    continue
                
                # Mark as processed
                processed_medicines[medicine_key] = unit_price
                
                # Check if medicine with this name exists anywhere in the database
                check_result = tenant_session.execute(
                    text("""
                        SELECT id, medicine_name, unit_price, expiry_date FROM medicines 
                        WHERE LOWER(TRIM(medicine_name)) = LOWER(TRIM(:medicine_name))
                        AND price_list_id IN (
                            SELECT id FROM price_lists WHERE tenant_id = :tenant_id
                        )
                        LIMIT 1
                    """),
                    {
                        'medicine_name': medicine_name,
                        'tenant_id': tenant_id
                    }
                )
                existing_medicine = check_result.fetchone()
                
                if existing_medicine:
                    # Medicine exists - update ALL instances of this medicine name with new price and expiry_date
                    # Use exact same comparison as check query (TRIM on both sides)
                    old_price = float(existing_medicine.unit_price) if existing_medicine.unit_price else 0
                    update_result = tenant_session.execute(
                        text("""
                            UPDATE medicines
                            SET unit_price = :unit_price,
                                expiry_date = :expiry_date,
                                last_updated = CURRENT_TIMESTAMP,
                                is_active = TRUE
                            WHERE LOWER(TRIM(medicine_name)) = LOWER(TRIM(:medicine_name))
                            AND price_list_id IN (
                                SELECT id FROM price_lists WHERE tenant_id = :tenant_id
                            )
                        """),
                        {
                            'medicine_name': medicine_name,
                            'unit_price': unit_price,
                            'expiry_date': expiry_date,
                            'tenant_id': tenant_id
                        }
                    )
                    rows_updated = update_result.rowcount
                    if rows_updated > 0:
                        logger.info(f"Updated {rows_updated} instance(s) of medicine '{medicine_name}' from price {old_price} to {unit_price}, expiry_date: {expiry_date}")
                    else:
                        logger.warning(f"Medicine '{medicine_name}' found but UPDATE affected 0 rows")
                else:
                    # Medicine doesn't exist - insert new
                    tenant_session.execute(
                        text("""
                            INSERT INTO medicines (price_list_id, medicine_name, unit_price, expiry_date)
                            VALUES (:price_list_id, :medicine_name, :unit_price, :expiry_date)
                        """),
                        {
                            'price_list_id': price_list_id,
                            'medicine_name': medicine_name,
                            'unit_price': unit_price,
                            'expiry_date': expiry_date
                        }
                    )
                    logger.info(f"Inserted new medicine '{medicine_name}' with price {unit_price}, expiry_date: {expiry_date}")
            
            # Activate price list
            tenant_session.execute(
                text("""
                    UPDATE price_lists 
                    SET status = 'active', activated_at = NOW()
                    WHERE id = :id
                """),
                {'id': price_list_id}
            )
            
            tenant_session.commit()
            tenant_session.close()
            
            return jsonify({
                'success': True,
                'statistics': {
                    'total_items': len(valid_records),
                    'valid_items': len(valid_records),
                    'invalid_items': 0
                }
            }), 201
            
        except Exception as e:
            tenant_session.rollback()
            tenant_session.close()
            logger.error(f"Database error during manual upload: {e}")
            raise
        
    except Exception as e:
        error_details = traceback.format_exc()
        logger.error(f"Manual upload error: {e}")
        logger.error(f"Full traceback:\n{error_details}")
        return jsonify({
            'error': 'Upload failed', 
            'message': str(e),
            'error_type': type(e).__name__
        }), 500


@depot_bp.route('/download-prices-manual', methods=['POST'])
@require_auth
@require_role('depot')
def download_prices_manual():
    """Get prices for manually entered medicine names - with flexible matching (ignores numbers)"""
    try:
        tenant_id = request.current_user.get('tenant_id')
        if not tenant_id:
            return jsonify({'error': 'User not associated with a tenant'}), 400
        
        # Get medicine names from request
        data = request.get_json()
        if not data or 'medicine_names' not in data:
            return jsonify({'error': 'No medicine names provided'}), 400
        
        medicine_names = data['medicine_names']
        if not isinstance(medicine_names, list) or len(medicine_names) == 0:
            return jsonify({'error': 'No valid medicine names provided'}), 400
        
        # Get tenant database name
        session = CentralDB.get_session()
        tenant_result = session.execute(
            text("SELECT database_name FROM tenants WHERE id = :tenant_id"),
            {'tenant_id': tenant_id}
        )
        tenant = tenant_result.fetchone()
        if not tenant:
            session.close()
            return jsonify({'error': 'Tenant not found'}), 404
        
        tenant_db_name = tenant.database_name
        session.close()
        
        # Query medicines from tenant database with flexible matching
        tenant_session = TenantDBManager.get_session(tenant_db_name)
        
        matched_medicines = []
        not_found = []
        
        for name in medicine_names:
            try:
                name_clean = name.strip()
                if not name_clean:
                    continue
                
                # First try exact match (case-insensitive)
                db_result = tenant_session.execute(
                    text("""
                        SELECT m.id, m.medicine_name, m.unit_price, m.expiry_date
                        FROM medicines m
                        INNER JOIN price_lists pl ON m.price_list_id = pl.id
                        WHERE pl.tenant_id = :tenant_id 
                            AND LOWER(TRIM(m.medicine_name)) = LOWER(TRIM(:name))
                            AND m.is_active = TRUE
                        ORDER BY pl.activated_at DESC
                        LIMIT 1
                    """),
                    {'tenant_id': tenant_id, 'name': name_clean}
                )
                medicine = db_result.fetchone()
                
                # If no exact match, try matching without numbers
                if not medicine:
                    # Remove numbers and normalize for matching
                    name_normalized = re.sub(r'\d+', '', name_clean.lower().strip())
                    name_normalized = ' '.join(name_normalized.split())  # Normalize spaces
                    
                    # Get all medicines and match manually
                    all_medicines_result = tenant_session.execute(
                        text("""
                            SELECT m.id, m.medicine_name, m.unit_price, m.expiry_date, pl.activated_at
                            FROM medicines m
                            INNER JOIN price_lists pl ON m.price_list_id = pl.id
                            WHERE pl.tenant_id = :tenant_id 
                                AND m.is_active = TRUE
                            ORDER BY pl.activated_at DESC
                        """),
                        {'tenant_id': tenant_id}
                    )
                    all_medicines = all_medicines_result.fetchall()
                    
                    # Find best match (medicine name without numbers matches)
                    best_match = None
                    for med in all_medicines:
                        med_name_normalized = re.sub(r'\d+', '', med.medicine_name.lower().strip())
                        med_name_normalized = ' '.join(med_name_normalized.split())
                        
                        # Check if normalized names match
                        if name_normalized == med_name_normalized:
                            best_match = med
                            break
                        # Also check if input is contained in medicine name (without numbers)
                        elif name_normalized in med_name_normalized or med_name_normalized in name_normalized:
                            if not best_match:  # Take first match
                                best_match = med
                    
                    if best_match:
                        medicine = best_match
                
                if medicine:
                    expiry_date_str = medicine.expiry_date.isoformat() if medicine.expiry_date else None
                    matched_medicines.append({
                        'id': str(medicine.id),
                        'medicine_name': medicine.medicine_name,
                        'unit_price': float(medicine.unit_price),
                        'expiry_date': expiry_date_str
                    })
                else:
                    not_found.append(name_clean)
            except Exception as e:
                logger.error(f"Error querying medicine '{name}': {e}")
                not_found.append(name)
        
        tenant_session.close()
        
        return jsonify({
            'success': True,
            'matched_medicines': matched_medicines,
            'not_found': not_found,
            'total_requested': len(medicine_names),
            'total_matched': len(matched_medicines),
            'total_not_found': len(not_found)
        }), 200
        
    except Exception as e:
        logger.error(f"Manual download prices error: {e}", exc_info=True)
        return jsonify({'error': 'Failed to get prices', 'message': str(e)}), 500


@depot_bp.route('/download-prices-pdf', methods=['POST'])
@require_auth
@require_role('depot')
def download_prices_pdf():
    """Generate and download PDF file with medicine prices"""
    try:
        from flask import send_file
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import inch
        import io
        
        data = request.get_json()
        medicines = data.get('medicines', [])
        
        if not medicines:
            return jsonify({'error': 'No medicines data provided'}), 400
        
        # Create PDF in memory
        pdf_file = io.BytesIO()
        doc = SimpleDocTemplate(pdf_file, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        title = Paragraph("Medicine Price List", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Prepare table data
        table_data = [['Medicine Name', 'Unit Price (RWF)']]
        for med in medicines:
            # Format price as integer (remove .00)
            price = med.get('unit_price', 0)
            price_formatted = f"{int(round(price))}"
            table_data.append([
                med.get('medicine_name', ''),
                price_formatted
            ])
        
        # Create table
        table = Table(table_data, colWidths=[4.5*inch, 1.5*inch])
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f9ff')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        pdf_file.seek(0)
        
        return send_file(
            pdf_file,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'medicine_prices_{datetime.now().strftime("%Y%m%d")}.pdf'
        )
        
    except Exception as e:
        logger.error(f"PDF download error: {e}", exc_info=True)
        return jsonify({'error': 'Failed to generate PDF file', 'message': str(e)}), 500
